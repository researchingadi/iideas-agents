"""
Comparison pipeline.
Runs three archetype identification approaches at 10 % data increments
and measures agreement between them.

Approaches
----------
1. Old Script      — rule-based (ERA × SIZE × CONSTRUCTION)
2. AI Agent        — GPT-4o + KMeans (fresh session each increment)
3. ML Model        — Random Forest trained on old-script labels as ground truth

Output
------
comparison_results.json  — full structured results
comparison_results.xlsx  — formatted Excel report
"""

import warnings
warnings.filterwarnings("ignore")

import json
import re
import os

import numpy as np
import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from sklearn.cluster import KMeans
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import adjusted_rand_score, silhouette_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler

from run_old_script import run_old_classification

load_dotenv()


# ---------------------------------------------------------------------------
# Approach 1 – Old Script (imported from run_old_script.py)
# ---------------------------------------------------------------------------
# run_old_classification(df) is imported above.


# ---------------------------------------------------------------------------
# Approach 2 – AI Agent (GPT-4o, fresh session)
# ---------------------------------------------------------------------------
def run_agent_fresh(subset_df: pd.DataFrame) -> dict:
    """
    Run a fresh GPT-4o + KMeans session on subset_df.
    Each call is stateless — no memory of previous increments.
    """
    client = OpenAI(
        base_url="https://models.inference.ai.azure.com",
        api_key=os.getenv("GITHUB_TOKEN"),
    )

    sample   = subset_df.head(5).to_string()
    columns  = list(subset_df.columns)
    n_sample = len(subset_df)

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {
                    "role": "system",
                    "content": "You are a building scientist. Return only raw valid JSON, no markdown.",
                },
                {
                    "role": "user",
                    "content": (
                        f"Classify these {n_sample} buildings into archetypes.\n"
                        f"Available columns: {columns}\n"
                        f"Sample rows:\n{sample}\n\n"
                        'Return JSON only:\n'
                        '{"columns_to_use": [...], "n_clusters": <int 2-12>, '
                        '"archetype_names": [...]}'
                    ),
                },
            ],
        )
        raw = response.choices[0].message.content.strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
        raw = re.sub(r"```\s*$", "", raw, flags=re.MULTILINE)
        config = json.loads(raw.strip())
    except Exception as exc:
        print(f"  ⚠️  GPT-4o error: {exc} — using fallback config")
        config = {
            "columns_to_use": ["yearOfConstruction", "finishAreaSqft",
                                "constructionType", "numOfFloors"],
            "n_clusters": 6,
            "archetype_names": [f"Archetype {i}" for i in range(1, 7)],
        }

    # Filter to columns that exist
    cols_to_use = [c for c in config["columns_to_use"] if c in subset_df.columns]
    if not cols_to_use:
        cols_to_use = [c for c in
                       ["yearOfConstruction", "finishAreaSqft", "constructionType"]
                       if c in subset_df.columns]

    # Cap clusters so KMeans never gets more clusters than samples
    n_clusters = max(2, min(int(config.get("n_clusters", 6)), n_sample // 3))
    archetype_names = config.get("archetype_names", [])
    if len(archetype_names) != n_clusters:
        archetype_names = [f"Archetype {i+1}" for i in range(n_clusters)]

    # Encode features (same pattern as agent1_archetype.py)
    cluster_data = subset_df[cols_to_use].copy()
    for col in cluster_data.columns:
        numeric = pd.to_numeric(cluster_data[col], errors="coerce")
        if numeric.isna().sum() > cluster_data[col].isna().sum():
            cluster_data[col] = (
                pd.Categorical(cluster_data[col].fillna("Unknown").astype(str))
                .codes.astype(float)
            )
        else:
            cluster_data[col] = numeric
    cluster_data = cluster_data.fillna(0).astype(float)

    scaler = StandardScaler()
    scaled = scaler.fit_transform(cluster_data)

    kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
    labels = kmeans.fit_predict(scaled)

    sil = (silhouette_score(scaled, labels)
           if len(set(labels)) > 1 else 0.0)

    distribution = {archetype_names[i]: int((labels == i).sum())
                    for i in range(n_clusters)}

    return {
        "n_archetypes":     n_clusters,
        "distribution":     distribution,
        "silhouette_score": round(float(sil), 4),
        "labels":           labels,
    }


# ---------------------------------------------------------------------------
# Approach 3 – ML Model (Random Forest)
# ---------------------------------------------------------------------------
def run_ml_model(subset_df, agent_labels=None):
    """
    Trains a Random Forest classifier.
    Uses Agent 1 cluster labels as ground truth if provided,
    otherwise falls back to old script labels.
    This makes it a truly independent validation approach.
    """
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.model_selection import train_test_split
    from sklearn.metrics import accuracy_score

    # Use agent labels as ground truth if available
    if agent_labels is not None and len(agent_labels) == len(subset_df):
        y = pd.Series(agent_labels).astype(str)
        ground_truth_source = "AI Agent"
    else:
        labeled = run_old_classification(subset_df.copy())
        y = labeled["archetype_label"].fillna("Unknown").astype(str)
        ground_truth_source = "Old Script"

    # Drop invalid labels
    valid_mask = y.notna() & (y != "nan") & (y != "None") & (y != "-1")
    y          = y[valid_mask].reset_index(drop=True)
    subset_df  = subset_df[valid_mask].reset_index(drop=True)

    # Features - use numeric building attributes
    feature_cols = [
        "yearOfConstruction", "finishAreaSqft",
        "numOfBedrooms", "numOfFloors", "sizeSqft",
        "age", "grossFloorArea", "numOfRoomsTotal",
        "buildingHeightAverage", "buildingAssessedValueTotal"
    ]
    feature_cols = [c for c in feature_cols if c in subset_df.columns]

    # Build X - force everything to numeric
    X = pd.DataFrame()
    for col in feature_cols:
        X[col] = pd.to_numeric(subset_df[col], errors="coerce").fillna(0)
    X = X.fillna(0).astype(float)

    n_classes = y.nunique()
    n_samples = len(X)

    if n_samples < 15 or n_classes < 2:
        return {
            "n_archetypes":     n_classes,
            "distribution":     y.value_counts().to_dict(),
            "silhouette_score": "N/A (supervised)",
            "accuracy":         f"N/A ({n_samples} buildings / {n_classes} classes)",
            "ground_truth":     ground_truth_source,
            "labels":           y.values,
            "feature_importance": {},
        }

    # Reduce classes if too many relative to sample count
    max_classes = max(2, n_samples // 5)
    if n_classes > max_classes:
        top_classes = y.value_counts().head(max_classes).index
        mask      = y.isin(top_classes)
        X         = X[mask].reset_index(drop=True)
        y         = y[mask].reset_index(drop=True)
        n_classes = y.nunique()

    # 80/20 stratified split, plain split as fallback
    try:
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42, stratify=y
        )
    except ValueError:
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42
        )

    rf = RandomForestClassifier(
        n_estimators=100,
        random_state=42,
        min_samples_leaf=1,
    )
    rf.fit(X_train, y_train)

    y_pred    = rf.predict(X_test)
    accuracy  = accuracy_score(y_test, y_pred)
    all_preds = rf.predict(X)

    importance   = dict(zip(feature_cols,
                            [round(float(v), 4) for v in rf.feature_importances_]))
    top_features = dict(sorted(importance.items(),
                               key=lambda x: x[1], reverse=True)[:3])

    return {
        "n_archetypes":     n_classes,
        "distribution":     pd.Series(all_preds).value_counts().to_dict(),
        "silhouette_score": "N/A (supervised)",
        "accuracy":         round(accuracy, 4),
        "ground_truth":     ground_truth_source,
        "labels":           all_preds,
        "feature_importance": top_features,
    }


# ---------------------------------------------------------------------------
# Agreement metric
# ---------------------------------------------------------------------------
def calculate_agreement(labels1, labels2) -> float | str:
    """Adjusted Rand Index between two label arrays (any comparable type)."""
    try:
        l1 = list(labels1)
        l2 = list(labels2)
        if len(l1) == len(l2) and len(l1) > 1:
            return round(float(adjusted_rand_score(l1, l2)), 4)
    except Exception:
        pass
    return "N/A"


# ---------------------------------------------------------------------------
# Console table
# ---------------------------------------------------------------------------
def print_comparison_table(comparison: dict) -> None:
    pct = comparison["increment_pct"]
    n   = comparison["n_buildings"]

    print(f"\n{'─' * 62}")
    print(f"  RESULTS AT {pct}%  ({n} buildings)")
    print(f"{'─' * 62}")
    print(f"  {'Approach':<20} {'Archetypes':>11} {'Silhouette':>12} {'Accuracy':>10}")
    print(f"  {'─' * 58}")

    old = comparison["old_script"]
    agt = comparison["ai_agent"]
    ml  = comparison["ml_model"]

    print(f"  {'Old Script':<20} {old['n_archetypes']:>11} {'N/A':>12} {'N/A':>10}")
    print(f"  {'AI Agent':<20} {agt['n_archetypes']:>11} "
          f"{str(agt['silhouette_score']):>12} {'N/A':>10}")
    print(f"  {'ML Model':<20} {ml['n_archetypes']:>11} "
          f"{'N/A':>12} {str(ml['accuracy']):>10}")
    if comparison["ml_model"].get("feature_importance"):
        top = list(comparison["ml_model"]["feature_importance"].keys())[:3]
        print(f"  ML Top Features: {', '.join(top)}")
    print(f"  ML Ground Truth: {comparison['ml_model'].get('ground_truth', 'N/A')}")

    ag = comparison["agreement"]
    print(f"\n  Agreement (Adjusted Rand Index):")
    print(f"    Old vs Agent : {ag['old_vs_agent_ari']}")
    print(f"    Old vs ML    : {ag['old_vs_ml_ari']}")
    print(f"    Agent vs ML  : {ag['agent_vs_ml_ari']}")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def save_comparison_excel(results: list[dict]) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    headers = [
        "Increment %", "Buildings",
        "Old Script — Archetypes",
        "Agent — Archetypes", "Agent — Silhouette",
        "ML — Archetypes", "ML — Accuracy",
        "ARI: Old vs Agent", "ARI: Old vs ML", "ARI: Agent vs ML",
    ]
    header_fill = PatternFill(patternType="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    center      = Alignment(horizontal="center")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 2, 14)

    for ri, r in enumerate(results, 2):
        row_vals = [
            r["increment_pct"],
            r["n_buildings"],
            r["old_script"]["n_archetypes"],
            r["ai_agent"]["n_archetypes"],
            r["ai_agent"]["silhouette_score"],
            r["ml_model"]["n_archetypes"],
            str(r["ml_model"]["accuracy"]),
            str(r["agreement"]["old_vs_agent_ari"]),
            str(r["agreement"]["old_vs_ml_ari"]),
            str(r["agreement"]["agent_vs_ml_ari"]),
        ]
        for ci, val in enumerate(row_vals, 1):
            ws.cell(row=ri, column=ci, value=val).alignment = center

    wb.save("comparison_results.xlsx")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------
def run_comparison(csv_path: str) -> list[dict]:
    df    = pd.read_csv(csv_path)
    total = len(df)

    print(f"\n{'=' * 62}")
    print(f"  IIDEAS LAB — COMPARISON PIPELINE")
    print(f"  Dataset: {os.path.basename(csv_path)}  ({total} buildings)")
    print(f"  Approaches: Old Script | AI Agent (GPT-4o) | ML (RF)")
    print(f"{'=' * 62}")

    increments = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
    results: list[dict] = []

    for pct in increments:
        n      = max(5, int(total * pct))
        subset = df.head(n).copy()

        print(f"\n{'=' * 62}")
        print(f"  TESTING ON {int(pct * 100)}% OF DATA  ({n} buildings)")
        print(f"{'=' * 62}")

        # ── Approach 1: Old Script ────────────────────────────────────────
        print("\n  [1/3] Running old rule-based script...")
        old_result = run_old_classification(subset.copy())
        old_labels = old_result["archetype_label"].values

        # ── Approach 2: AI Agent ──────────────────────────────────────────
        print("  [2/3] Running AI Agent (GPT-4o, fresh session)...")
        agent_result = run_agent_fresh(subset.copy())

        # ── Approach 3: ML Model ─────────────────────────────────────────
        print("  [3/3] Running ML Model (Random Forest)...")
        agent_labels = agent_result.get("labels", None)
        ml_result = run_ml_model(subset.copy(), agent_labels)

        # ── Agreement ────────────────────────────────────────────────────
        comparison = {
            "increment_pct": int(pct * 100),
            "n_buildings":   n,

            "old_script": {
                "n_archetypes": int(old_result["archetype_label"].nunique()),
                "distribution": old_result["archetype_label"]
                                          .value_counts().to_dict(),
                "method": "Rule-based (Era × Size × Construction)",
            },

            "ai_agent": {
                "n_archetypes":     agent_result["n_archetypes"],
                "distribution":     agent_result["distribution"],
                "silhouette_score": agent_result["silhouette_score"],
                "method": "GPT-4o + KMeans clustering",
            },

            "ml_model": {
                "n_archetypes":     ml_result["n_archetypes"],
                "distribution":     ml_result["distribution"],
                "silhouette_score": str(ml_result["silhouette_score"]),
                "accuracy":         str(ml_result["accuracy"]),
                "ground_truth":     ml_result.get("ground_truth", "N/A"),
                "feature_importance": ml_result.get("feature_importance", {}),
                "method": "Random Forest (trained on Agent labels)",
            },

            "agreement": {
                "old_vs_agent_ari": calculate_agreement(
                    old_labels, agent_result["labels"]),
                "old_vs_ml_ari":    calculate_agreement(
                    old_labels, ml_result["labels"]),
                "agent_vs_ml_ari":  calculate_agreement(
                    agent_result["labels"], ml_result["labels"]),
            },
        }

        results.append(comparison)
        print_comparison_table(comparison)

    # ── Save outputs ──────────────────────────────────────────────────────────
    with open("comparison_results.json", "w") as f:
        json.dump(results, f, indent=2, default=str)

    save_comparison_excel(results)

    print(f"\n{'=' * 62}")
    print("  ✅ Comparison complete!")
    print("     comparison_results.json  — full structured results")
    print("     comparison_results.xlsx  — formatted Excel report")
    print(f"{'=' * 62}\n")

    return results


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "community_data.csv"
    run_comparison(path)
